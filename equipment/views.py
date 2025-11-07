from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Q
from django.http import JsonResponse
from django.db import IntegrityError
from rest_framework import viewsets, filters, status
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.pagination import PageNumberPagination
from django_filters.rest_framework import DjangoFilterBackend
from django.urls import reverse
from .models import Equipment, Maintenance, Notification
from .serializers import (
    EquipmentSerializer, MaintenanceSerializer,
    NotificationSerializer
)
from django.views.decorators.http import require_http_methods
from django.contrib.auth.decorators import login_required

from django.views.decorators.csrf import csrf_exempt
import json
from django.contrib.auth.models import Group
from django import forms
from django.utils.dateparse import parse_date
from django.views.generic import ListView
from django.shortcuts import  get_object_or_404
from django.contrib import messages
from django.contrib.auth.views import LoginView
from django.utils.translation import gettext as _
from django.http import JsonResponse 
from django.contrib.auth import logout
from django.urls import reverse_lazy
from django.views import View
from django.contrib import messages
from django.shortcuts import render, redirect
from django.shortcuts import render, redirect
from django.contrib.auth import logout
from django.contrib import messages
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from .models import CustomUser


from django.contrib.auth import logout
from django.shortcuts import redirect

def logout_view(request):
    logout(request)
    return redirect('login')  # أو الصفحة التي تريد أن يعود إليها المستخدم

from django.shortcuts import render, redirect
from django.views.generic import TemplateView
from .forms import CustomUserForm  # إذا كان لديك نموذج مخصص لإضافة مستخدم
from django.contrib.auth import get_user_model
from django.views.generic import ListView

class user_list(TemplateView):
    template_name = 'user_list.html'
    
    def get(self, request, *args, **kwargs):
        # استرجاع قائمة المستخدمين من قاعدة البيانات
        users = CustomUser.objects.all()
        form = CustomUserForm()  # إنشاء نموذج لإضافة مستخدم جديد
        return self.render_to_response({'users': users, 'form': form})

    def post(self, request, *args, **kwargs):
        form = CustomUserForm(request.POST)
        if form.is_valid():
            form.save()  # حفظ المستخدم الجديد
            return redirect('user_list')  # إعادة التوجيه إلى صفحة المستخدمين
        users = CustomUser.objects.all()
        return self.render_to_response({'users': users, 'form': form})

from django.views.generic import TemplateView
from django.contrib.auth import get_user_model

class UserListAndAddView(TemplateView):
    template_name = 'user_list.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        User = get_user_model()
        context['users'] = User.objects.all()
        return context
# views.py
from django.shortcuts import render
from django.views.generic import ListView
from .models import CustomUser

class UserListView(ListView):
    model = CustomUser
    template_name = 'user_list.html'  # تأكد من أن الملف موجود
    context_object_name = 'users'
from .models import Activite
from django.utils.timezone import now
# ✅ Vue du tableau de bord
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from .models import Equipment, Maintenance, Notification, Demande ,Activite
from .models import CustomUser  # تأكد من أن المسار صحيح حسب تطبيقك
from django.utils import timezone
@login_required
def dashboard(request):
    total_users = CustomUser.objects.count()
    total_equipments = Equipment.objects.count()

    in_maintenance = Maintenance.objects.filter(status='en_cours').count()
    total_pannes = Equipment.objects.filter(state='En Panne').count()

    notif_today = Notification.objects.filter(created_at__date__exact=datetime.date.today()).count()
    notif_unread = Notification.objects.filter(recipient_user=request.user, is_read=False).count()
    unread_count = Notification.objects.filter(recipient_user=request.user, is_read=False).count()

    open_requests = Demande.objects.filter(statut='ouverte').count()

    recent_activity = Activite.objects.order_by('-date')[:10]  # استرجاع آخر 10 أنشطة مثلا

    recent_notifications = Notification.objects.order_by('-created_at')[:5]


    context = {
        'current_user': request.user.get_full_name() if hasattr(request.user, 'get_full_name') else f"{request.user.first_name} {request.user.last_name}",
        'total_users': total_users,
        'total_equipments': total_equipments,
        'in_maintenance': in_maintenance,
        'unread_count': unread_count,
        'notif_today': notif_today,
        'open_requests': open_requests,
        'recent_activity': recent_activity,
        'recent_notifications': recent_notifications,
        'notif_unread': Notification.objects.filter(is_read=False, recipient_user=request.user).count(),
         
        'notif_unread': notif_unread,
        'total_pannes': total_pannes,
    }
    print("Total Pannes: ", total_pannes)

    return render(request, 'dashboard.html', context)
# views.py
from rest_framework import viewsets
from .models import Panne
from .serializers import PanneSerializer

class PanneViewSet(viewsets.ModelViewSet):
    queryset = Panne.objects.all().order_by('-date_signalement')
    serializer_class = PanneSerializer

from django.http import JsonResponse
from django.db.models import Count
from django.utils.timezone import now
from django.views.decorators.http import require_GET
import calendar

from django.http import JsonResponse
from django.db.models import Count
from django.utils.timezone import now
from django.views.decorators.http import require_GET
import calendar

@require_GET
def maintenance_data_api(request):
    # الحصول على السنة من بارامترات GET، وإذا لم تُرسل نأخذ السنة الحالية
    year = request.GET.get('year')
    if year is not None and year.isdigit():
        year = int(year)
    else:
        year = now().year

    # فلترة حسب تاريخ الانتهاء (date_fin) خلال السنة المختارة
    queryset = (
        Maintenance.objects.filter(date_fin__year=year)
        .values('date_fin__month')
        .annotate(total_count=Count('id'))
        .order_by('date_fin__month')
    )

    labels = []
    values = []
    for entry in queryset:
        month_num = entry['date_fin__month']
        month_name = calendar.month_abbr[month_num]
        labels.append(month_name)
        values.append(entry['total_count'])

    data = {
        "labels": labels,
        "values": values
    }

    return JsonResponse(data)
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required

@login_required
def notif_unread_count(request):
    count = Notification.objects.filter(user=request.user, is_read=False).count()
    return JsonResponse({'count': count})
from django.contrib.auth.decorators import login_required
from .models import Notification
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
import json
# عرض قائمة الإشعارات
@require_http_methods(["GET"])
def get_notifications(request):
    # إذا كان المستخدم غير مسجل الدخول، يمكنك تعديل الاستعلام حسب الحاجة
    notifications = Notification.objects.all().values('id', 'message', 'type', 'created_at', 'is_read')
    return JsonResponse(list(notifications), safe=False)
from django.views.decorators.csrf import csrf_exempt

@csrf_exempt
@require_http_methods(["POST"])
def add_notification(request):
    data = json.loads(request.body)
    message = data.get("message")
    type = data.get("type")
    user_email = data.get("user_email")
    all_users = data.get("all_users", False)

    from django.contrib.auth.models import User
    user = None
    if user_email:
        user = User.objects.filter(email=user_email).first()

    notification = Notification.objects.create(
        message=message,
        type=type,
        user=user,
        all_users=all_users
    )

    return JsonResponse({'success': True, 'notification_id': notification.id})


from django.views.decorators.http import require_http_methods

# حذف إشعار
@require_http_methods(["DELETE"])
def delete_notification(request, id):
    try:
        notification = Notification.objects.get(id=id)
        notification.delete()
        return JsonResponse({'success': True})
    except Notification.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Notification not found'}, status=404)
from django.shortcuts import render
from .models import Notification  # Assure-toi que le modèle Notification existe
from django.shortcuts import render
from .models import Notification
from django.db.models import Q

def notifications_view(request):
    notifications = Notification.objects.filter(recipient_user=request.user).order_by('-created_at')
    Notification.objects.filter(recipient_user=request.user, is_read=False).update(is_read=True)

    # جلب الإشعارات العامة أو الخاصة بهذا المستخدم

    return render(request, 'notifications.html', {'notifications': notifications})
from rest_framework.decorators import api_view
from rest_framework.response import Response
from .models import Notification
from .serializers import NotificationSerializer

@api_view(['GET'])
def user_notifications(request):
    user = request.user
    if user.is_authenticated:
        notifications = Notification.objects.filter(user=user).order_by('-created_at')
        serializer = NotificationSerializer(notifications, many=True)
        return Response(serializer.data)
    return Response([], status=401)

def my_view(request):
    utilisateurs_url = reverse('utilisateurs')
    return JsonResponse({'url': utilisateurs_url})
from django.contrib.auth import get_user_model
User = get_user_model()
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from .models import Notification
from .serializers import NotificationSerializer

class NotificationListAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        notifications = Notification.objects.all().order_by('-created_at')
        serializer = NotificationSerializer(notifications, many=True)
        return Response(serializer.data)
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.shortcuts import render
from django.utils.dateparse import parse_date
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.contrib.auth.models import Group
from .models import Equipment
from .serializers import GroupSerializer

import random
import string

def generate_password(length=10):
    chars = string.ascii_letters + string.digits
    return ''.join(random.choice(chars) for _ in range(length))
from django.contrib import messages
from .models import CustomUser
from .views import generate_password  # إذا وضعتها في ملف utils.py

def create_user_view(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        role = request.POST.get('role')
        is_active = request.POST.get('is_active') == 'on'

        password = generate_password()
        user = CustomUser(
            email=email,
            first_name=first_name,
            last_name=last_name,
            role=role,
            is_active=is_active
        )
        user.set_password(password)
        user.save()

        # ✅ عرض كلمة المرور مرة واحدة
        messages.success(request, f"Utilisateur créé. Mot de passe : {password}")
        return redirect('user_list')  # أو أي صفحة تريد

# دالة مساعدة لحساب الإحصائيات مع التواريخ
def statistiques(start_date=None, end_date=None):
    query = Equipment.objects.all()

    if start_date and end_date:
        query = query.filter(created_at__date__range=(start_date, end_date))

    total = query.count()
    en_panne = Equipment.objects.filter(state='En Panne').count()
    en_maintenance = query.filter(state='En maintenance').count()
    en_fonctionnement = query.filter(state='En fonctionnement').count()

    return {
        'total': total,
        'En_panne': en_panne,
        'En_maintenance': en_maintenance,
        'En_fonctionnement': en_fonctionnement,
    }


# عرض صفحة HTML للإحصائيات
@login_required
def statistiques_page_view(request):
    context = statistiques()
    return render(request, 'statistiques.html', context)


# API تُرجع الإحصائيات على شكل JSON
@login_required
def statistiques_api(request):
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')

    try:
        start_date = parse_date(start_date_str) if start_date_str else None
        end_date = parse_date(end_date_str) if end_date_str else None
    except ValueError:
        return JsonResponse({'error': 'Date invalide'}, status=400)

    data = statistiques(start_date, end_date)
    return JsonResponse(data)

from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import get_object_or_404, render
from django.contrib import messages
import json
from django.shortcuts import render, redirect
from django.contrib import messages
from .models import Maintenance, Equipment, CustomUser
from .forms import MaintenanceForm

def maintenance_view(request):
    maintenances = Maintenance.objects.all()
    equipments = Equipment.objects.all()

    # فقط المستخدمين الذين لديهم الدور "technicien"
    techs = CustomUser.objects.filter(role='technician')


    # استرجاع الفلاتر من GET
    machine = request.GET.get('machine')
    etat = request.GET.get('etat')
    tech = request.GET.get('tech')
    date = request.GET.get('date')
    urgence = request.GET.get('urgence')

    # تطبيق الفلاتر
    if machine:
        maintenances = maintenances.filter(machine__name__icontains=machine)
    if etat:
        maintenances = maintenances.filter(status=etat)
    if tech:
        maintenances = maintenances.filter(technicien__id=tech)
    if date:
        maintenances = maintenances.filter(date=date)
    if urgence:
        maintenances = maintenances.filter(urgence=urgence)

    # معالجة الفورم
    if request.method == "POST":
        form = MaintenanceForm(user=request.user)

        if form.is_valid():
            maintenance = form.save(commit=False)

            # إذا كان المستخدم تقني ولم يتم تعيين حقل "technicien" من الفورم (لأنه مخفي)
            if request.user.role == 'technicien':
                maintenance.technicien = request.user

            maintenance.save()
            messages.success(request, "✅ Maintenance ajoutée avec succès.")
            return redirect('maintenance_view')
        else:
            messages.error(request, "❌ Erreur dans le formulaire, veuillez vérifier les champs.")
    else:
        form = MaintenanceForm(user=request.user)

    # إرسال البيانات إلى القالب
    context = {
        'maintenances': maintenances,
        'form': form,
        'techs': techs,
        'equipments': equipments,
        'urgence': urgence,
    }

    return render(request, 'maintenance.html', context)

from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from django.utils.timezone import now
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.contrib.auth import get_user_model

from .models import Maintenance, Notification, Activite
from .forms import MaintenanceForm

@csrf_exempt
@login_required
def add_maintenance(request):
    try:
        if request.user.role not in ['admin', 'technicien']:
            return JsonResponse({"message": "Permission refusée."}, status=403)

        if request.method != 'POST':
            return JsonResponse({"message": "Méthode non autorisée."}, status=405)

        form = MaintenanceForm(request.POST, request.FILES, user=request.user)

        if not form.is_valid():
            return JsonResponse({
                "message": "Erreur dans le formulaire.",
                "errors": form.errors
            }, status=400)

        maintenance = form.save(commit=False)

        if request.user.role == 'technicien':
            maintenance.technicien = request.user

        maintenance.save()

        Activite.objects.create(
            utilisateur=request.user,
            description=f"Ajout d’une maintenance pour l’équipement : {maintenance.machine.name}",
            date=now()
        )

        return JsonResponse({"message": "Maintenance ajoutée avec succès !"}, status=201)

    except Exception as e:
        print("🔴 Erreur interne :", str(e))
        return JsonResponse({
            "message": "Erreur serveur interne.",
            "error": str(e)  # فقط للتصحيح، أزل هذا في الإنتاج
        }, status=500)

from django.shortcuts import get_object_or_404
from django.http import JsonResponse
from .models import Maintenance

def delete_maintenance(request, maintenance_id):
    maintenance = get_object_or_404(Maintenance, id=maintenance_id)

    if not request.user.is_authenticated:
        return JsonResponse({"message": "Non autorisé"}, status=401)

    # ✅ Seul le créateur (technicien) ou un admin peut supprimer
    if request.user != maintenance.technicien and request.user.role != 'admin':
        return JsonResponse({"message": "Permission refusée. Vous ne pouvez pas supprimer cette maintenance."}, status=403)

    maintenance.delete()
    return JsonResponse({'message': 'Maintenance supprimée avec succès !'})
from django.contrib import messages

def edit_maintenance(request, pk):
    maintenance = get_object_or_404(Maintenance, pk=pk)

    if not request.user.is_authenticated:
        messages.error(request, "Vous devez être connecté pour modifier une maintenance.")
        return redirect('maintenance')

    if request.user != maintenance.technicien and request.user.role != 'admin':
        messages.error(request, "⛔ Permission refusée. Vous ne pouvez pas modifier cette maintenance.")
        return redirect('maintenance')  # أو أي صفحة مناسبة

    if request.method == 'POST':
        form = MaintenanceForm(request.POST, request.FILES, instance=maintenance)
        if form.is_valid():
            form.save()
            messages.success(request, "✅ Maintenance modifiée avec succès.")
            return redirect('maintenance')
    else:
        form = MaintenanceForm(instance=maintenance)

    return render(request, 'edit_maintenance.html', {'form': form})


class StandardResultsSetPagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = "page_size"
    max_page_size = 100
from django.contrib import messages
from django.shortcuts import render, redirect
from .forms import EquipmentForm
from .models import Activite
from django.utils.timezone import now
import logging
from django.shortcuts import get_object_or_404
from django.shortcuts import get_object_or_404

@login_required
def delete_equipment(request, id):
    if request.user.role != 'admin':
        messages.error(request, "❌ Vous n'avez pas la permission de supprimer un équipement.")
        return redirect('equipment_list')

    equipment = get_object_or_404(Equipment, id=id)

    if request.method == 'POST':
        equipment.delete()
        messages.success(request, "✅ L'équipement a été supprimé avec succès.")
        return redirect('equipment_list')

    return redirect('equipment_list')


def equipment_detail(request, id):
    equipment = get_object_or_404(Equipment, id=id)
    return render(request, 'equipment_detail.html', {'equipment': equipment})

from django.shortcuts import render, redirect
from django.http import JsonResponse
from django.contrib import messages
from django.utils.timezone import now
from django.conf import settings
from django.contrib.auth.decorators import login_required
from .forms import EquipmentForm
from .models import Equipment, Activite  # تأكد أنك استوردت النموذج
import logging

logger = logging.getLogger(__name__)
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.shortcuts import render, redirect
from .forms import EquipmentForm
from .models import Equipment, Activite
from django.utils.timezone import now

@login_required
def equipment_list(request):
    form = EquipmentForm()
    
    if request.method == "POST":
        # ✅ تحقق من صلاحيات المستخدم
        if request.user.role != 'admin':
            messages.error(request, "❌ Vous n'avez pas la permission d’ajouter un équipement.")
            return redirect('equipment_list')
        
        form = EquipmentForm(request.POST, request.FILES)
        if form.is_valid():
            equipment = form.save()

            # 🔔 تسجل النشاط
            Activite.objects.create(
                utilisateur=request.user,
                description=f"Ajout d’un nouvel équipement : {equipment.name}",
                date=now()
            )

            messages.success(request, "✅ Équipement ajouté avec succès.")
            return redirect('equipment_list')
        else:
            messages.error(request, "❌ Erreur dans le formulaire.")

    equipments = Equipment.objects.all()
    return render(request, 'equipment_list.html', {
        'equipments': equipments,
        'form': form,
        'is_admin': request.user.role == 'admin'  # لإخفاء الزر في القالب
    })

def filter_equipment(request):
    filter_value = request.GET.get('filter')
    if filter_value:
        equipments = Equipment.objects.filter(state=filter_value)
    else:
        equipments = Equipment.objects.all()

    data = []
    for eq in equipments:
        data.append({
            'id': eq.id,
            'name': eq.name,
            'state': eq.state,
            'serial_number': eq.serial_number,
            'image_url': eq.image.url if eq.image else '',
        })
    return JsonResponse({'equipments': data})
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.contrib import messages

@login_required
def edit_equipment(request, id):
    if request.user.role != 'admin':
        messages.error(request, "❌ Vous n'avez pas la permission de modifier un équipement.")
        return redirect('equipment_list')

    equipment = get_object_or_404(Equipment, id=id)

    if request.method == 'POST':
        form = EquipmentForm(request.POST, request.FILES, instance=equipment)
        if form.is_valid():
            form.save()
            messages.success(request, "✅ Équipement modifié avec succès.")
            return redirect('equipment_list')
        else:
            messages.error(request, "❌ Erreur dans le formulaire.")
    else:
        form = EquipmentForm(instance=equipment)

    return render(request, 'equipment/edit_equipment.html', {
        'form': form,
        'equipment': equipment
    })
   
class BaseViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter, DjangoFilterBackend]
    pagination_class = StandardResultsSetPagination

# ✅ API: Ajouter un équipement
class AjouterEquipementAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        serializer = EquipmentSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({"success": True, "id": serializer.data["id"]}, status=status.HTTP_201_CREATED)
        return Response({"success": False, "errors": serializer.errors}, status=status.HTTP_400_BAD_REQUEST)
 
# ✅ Maintenance ViewSet
class MaintenanceViewSet(BaseViewSet):
    queryset = Maintenance.objects.select_related('equipment').all()
    serializer_class = MaintenanceSerializer
    search_fields = ['equipment__name', 'status']

# ✅ Notification ViewSet
class NotificationViewSet(BaseViewSet):
    queryset = Notification.objects.order_by('-created_at').all()
    serializer_class = NotificationSerializer
    ordering_fields = ['created_at']
from rest_framework import viewsets
from .serializers import EquipmentSerializer
from rest_framework import status
from .permissions import IsAdmin, ReadOnly
class EquipmentViewSet(viewsets.ModelViewSet):
    queryset = Equipment.objects.all()
    serializer_class = EquipmentSerializer

    def list(self, request):
        """Retourne la liste de tous les équipements"""
        equipements = Equipment.objects.all()
        serializer = EquipmentSerializer(equipements, many=True)
        return Response(serializer.data)

    def create(self, request):
        """Ajouter un nouvel équipement"""
        serializer = EquipmentSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def retrieve(self, request, pk=None):
        """Récupérer un équipement spécifique"""
        try:
            equipement = Equipment.objects.get(pk=pk)
            serializer = EquipmentSerializer(equipement)
            return Response(serializer.data)
        except Equipment.DoesNotExist:
            return Response({'message': 'Équipement non trouvé'}, status=status.HTTP_404_NOT_FOUND)

    def update(self, request, pk=None):
        """Mettre à jour un équipement existant"""
        try:
            equipement = Equipment.objects.get(pk=pk)
            serializer = EquipmentSerializer(equipement, data=request.data)
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Equipment.DoesNotExist:
            return Response({'message': 'Équipement non trouvé'}, status=status.HTTP_404_NOT_FOUND)

    def partial_update(self, request, pk=None):
        """Mettre à jour partiellement un équipement"""
        try:
            equipement = Equipment.objects.get(pk=pk)
            serializer = EquipmentSerializer(equipement, data=request.data, partial=True)
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Equipment.DoesNotExist:
            return Response({'message': 'Équipement non trouvé'}, status=status.HTTP_404_NOT_FOUND)

    def destroy(self, request, pk=None):
        print(">> DESTROY METHOD CALLED BY:", request.user.email, "ROLE:", request.user.role)
 
        if not request.user.is_authenticated or request.user.role != 'admin':
           return Response({'message': '⚠️ Permission refusée.'}, status=status.HTTP_403_FORBIDDEN)

        try:
           equipement = Equipment.objects.get(pk=pk)
           equipement.delete()
           return Response({'message': '✅ Équipement supprimé avec succès!'}, status=status.HTTP_204_NO_CONTENT)
        except Equipment.DoesNotExist:
           return Response({'message': 'Équipement non trouvé'}, status=status.HTTP_404_NOT_FOUND)

    def get_permissions(self):
        if self.request.method in ['POST', 'PUT', 'PATCH', 'DELETE']:
            return [IsAdmin()]
        return [ReadOnly()]
# ✅ views.py
from django.contrib.auth import login, update_session_auth_hash
from django.contrib.auth.forms import AuthenticationForm, PasswordChangeForm
from django.contrib.auth.views import LoginView
from django.shortcuts import redirect, render
from django.views.generic import FormView

# ✅ Vue personnalisée de login avec redirection selon "must_change_password"
class CustomLoginView(LoginView):
    template_name = 'registration/login.html'  # Utilisez un template adapté
    authentication_form = AuthenticationForm

    def form_valid(self, form):
        user = form.get_user()
        login(self.request, user)
        if user.must_change_password:
            return redirect('force_password_change')
        return redirect('dashboard')  # Redirige vers la page principale


# ✅ Vue qui force l'utilisateur à changer son mot de passe
class ForcePasswordChangeView(FormView):
    template_name = 'registration/force_password_change.html'
    form_class = PasswordChangeForm

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('login')
        if not request.user.must_change_password:
            return redirect('dashboard')
        return super().dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        user = form.save()
        user.must_change_password = False  # ✅ Plus besoin de changer le mot de passe
        user.save()
        update_session_auth_hash(self.request, user)
        return redirect('dashboard')
from django.utils import translation

def view(request):
    user_language = 'fr'
    translation.activate(user_language)
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .forms import CustomUserForm
from .models import CustomUser


def is_admin(user):
    return user.is_authenticated and user.role == 'admin'
from functools import wraps

def admin_only(view_func):
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if request.user.is_authenticated:
            if request.user.role == 'admin':
                return view_func(request, *args, **kwargs)
            else:
                messages.error(request, "Vous n'avez pas la permission d'accéder à cette page.")
                return redirect('/dashboard/')
        else:
            return redirect('/login/')
    return wrapper

    return wrapper
from django.utils.timezone import now
from .models import CustomUser, Activite  # تأكد أن النموذج مستورد

@login_required
@admin_only
def manage_users(request):
    users = CustomUser.objects.all()
    edit_user = None

    if request.method == 'POST':
        if 'user_id' in request.POST:
            user = get_object_or_404(CustomUser, id=request.POST['user_id'])
            form = CustomUserForm(request.POST, instance=user)
            action = "modifié"
        else:
            form = CustomUserForm(request.POST)
            action = "ajouté"

        if form.is_valid():
            saved_user = form.save()

            # 🔔 تسجيل النشاط
            Activite.objects.create(
                utilisateur=request.user,
                description=f"Utilisateur {saved_user.email} {action} par admin.",
                date=now()
            )

            return redirect('user_list')

    elif 'edit_id' in request.GET:
        edit_user = get_object_or_404(CustomUser, id=request.GET['edit_id'])
        form = CustomUserForm(instance=edit_user)

    elif 'delete_id' in request.GET:
        user = get_object_or_404(CustomUser, id=request.GET['delete_id'])

        # 🔔 تسجيل النشاط قبل الحذف
        Activite.objects.create(
            utilisateur=request.user,
            description=f"Utilisateur {user.email} supprimé par admin.",
            date=now()
        )

        user.delete()
        return redirect('user_list')

    else:
        form = CustomUserForm()

    return render(request, 'user_list.html', {
        'users': users,
        'form': form,
        'edit_user': edit_user,
    })

# ✅ API لحذف المستخدم
class UserDeleteAPIView(APIView):
    def delete(self, request, id):
        if not request.user.is_authenticated or not request.user.is_superuser:
            return Response({"error": "Accès refusé"}, status=status.HTTP_403_FORBIDDEN)
        try:
            user = CustomUser.objects.get(id=id)
        except CustomUser.DoesNotExist:
            return Response({"error": "Utilisateur non trouvé"}, status=status.HTTP_404_NOT_FOUND)

        user.delete()
        return Response({"message": "Utilisateur supprimé avec succès"}, status=status.HTTP_204_NO_CONTENT)

from django.utils.timezone import now
from .models import Activite

def add_equipment(request):
    if request.method == 'POST':
        form = EquipmentForm(request.POST)
        if form.is_valid():
            equipment = form.save()

            # ✅ تسجيل النشاط
            if request.user.is_authenticated:
                Activite.objects.create(
                    utilisateur=request.user,
                    description=f"Ajout d’un nouvel équipement : {equipment.nom}",
                    date=now()
                )

            return redirect('equipment_list')  # Redirige vers la liste après ajout
    else:
        form = EquipmentForm()
    
    return render(request, 'equipment_list.html', {'form': form})
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from .models import Panne, Equipment
from .forms import PanneForm

from .models import Equipment
from .forms import PanneForm

def pannes_view(request):
    form = PanneForm()
    equipements = Equipment.objects.all()
    pannes = Panne.objects.all().order_by('-date_signalement')  # مثلًا

    return render(request, 'pannes.html', {
        'form': form,
        'equipements': equipements,
        'pannes': pannes
    })



# ✅ API pour créer une nouvelle panne
@csrf_exempt  
@require_POST
@login_required
def create_panne_api(request):
    try:
        form = PanneForm(request.POST, request.FILES)
        if form.is_valid():
            panne = form.save(commit=False)
            panne.signalee_par = request.user
            panne.save()

            return JsonResponse({'message': 'Panne enregistrée avec succès !', 'id': panne.id}, status=201)
        else:
            return JsonResponse({'errors': form.errors}, status=400)

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

from rest_framework.permissions import IsAuthenticated
from django.shortcuts import render, redirect
from .forms import PanneForm
from .models import Panne

def create_panne(request):
    if request.method == 'POST':
        form = PanneForm(request.POST, request.FILES)
        if form.is_valid():
            panne = form.save(commit=False)
            panne.signalee_par = request.user  # رابط المستخدم الحالي
            panne.save()
            return redirect('home')  # أو إلى صفحة قائمة الأعطال مثلاً
    else:
        form = PanneForm()

    return render(request, 'pannes.html', {'form': form})

from django.shortcuts import render, get_object_or_404
from .models import Panne

def detail_panne(request, id):
    panne = get_object_or_404(Panne, id=id)
    return render(request, 'detail_panne.html', {'panne': panne})
from django.shortcuts import render, get_object_or_404, redirect
from .models import Panne
from .forms import PanneForm  # Assure-toi que ce formulaire existe

def modifier_panne(request, id):
    panne = get_object_or_404(Panne, id=id)
    
    if request.method == 'POST':
        form = PanneForm(request.POST, request.FILES, instance=panne)
        if form.is_valid():
            form.save()
            return redirect('detail_panne', id=panne.id)
    else:
        form = PanneForm(instance=panne)

    return render(request, 'modifier_panne.html', {'form': form, 'panne': panne})

@login_required
def get_panne(request, id):
    if request.method == 'GET':
        try:
            panne = Panne.objects.select_related('equipment').get(id=id)
            return JsonResponse({
                "id": panne.id,
                "equipment_name": panne.equipment.name,
                "description": panne.description,
                "niveau_urgence": panne.niveau_urgence,
                "est_resolue": panne.est_resolue,
                "image": panne.image.url if panne.image else None,
            })
        except Panne.DoesNotExist:
            return JsonResponse({'error': 'Panne non trouvée'}, status=404)
@csrf_exempt
def update_panne(request, id):
    if request.method in ['PATCH', 'PUT']:
        try:
            panne = Panne.objects.get(id=id)
            data = json.loads(request.body)

            panne.description = data.get('description', panne.description)
            panne.niveau_urgence = data.get('niveau_urgence', panne.niveau_urgence)

            # Force conversion booléenne
            est_resolue_value = data.get('est_resolue')
            if isinstance(est_resolue_value, str):
                panne.est_resolue = est_resolue_value.lower() in ['true', '1', 'on']
            else:
                panne.est_resolue = bool(est_resolue_value)

            panne.save()

            return JsonResponse({'message': 'Panne mise à jour avec succès'})
        except Panne.DoesNotExist:
            return JsonResponse({'error': 'Panne non trouvée'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

@csrf_exempt
@login_required
def delete_panne(request, id):
    if request.method == 'DELETE':
        try:
            panne = Panne.objects.get(id=id)
            panne.delete()
            return JsonResponse({'message': 'Panne supprimée avec succès'})
        except Panne.DoesNotExist:
            return JsonResponse({'error': 'Panne non trouvée'}, status=404)
    return HttpResponseNotAllowed(['DELETE'])
from rest_framework.response import Response
from rest_framework.decorators import api_view
from .models import Panne
from .serializers import PanneSerializer

@api_view(['GET'])
def panne_detail(request, id):
    try:
        panne = Panne.objects.get(pk=id)
        serializer = PanneSerializer(panne)
        return Response(serializer.data)
    except Panne.DoesNotExist:
        return Response({'error': 'Panne introuvable'}, status=404)
    
from django.http import JsonResponse
from .models import Panne

def panne_detail(request, id):
    try:
        panne = Panne.objects.get(pk=id)
        data = {
            'equipment_name': panne.equipment.name,
            'description': panne.description,
            'niveau_urgence': panne.niveau_urgence,
            'est_resolue': panne.est_resolue,
            'image': panne.image.url if panne.image else None,
        }
        return JsonResponse(data)
    except Panne.DoesNotExist:
        return JsonResponse({'error': 'Panne introuvable'}, status=404)

import random
from django.utils import timezone
from datetime import timedelta
from django.contrib import messages
from django.contrib.auth import get_user_model
from django.core.mail import send_mail
from django.shortcuts import render, redirect
from .forms import VerificationCodeForm, SetNewPasswordForm
from django.shortcuts import render
from django.contrib.auth import get_user_model
from django.core.mail import send_mail
import random

from django.utils import timezone

def password_reset_request(request):
    context = {}
    if request.method == 'POST':
        email = request.POST.get('email')
        if not User.objects.filter(email=email).exists():
            context['error'] = "Cet e-mail n'existe pas dans notre base."
            return render(request, 'auth/password_reset_request.html', context)
        else:
            code = str(random.randint(100000, 999999))
            request.session['password_reset_code'] = code
            request.session['password_reset_email'] = email
            
            # حفظ وقت إنشاء الكود كـ ISO format string
            request.session['password_reset_code_time'] = timezone.now().isoformat()

            send_mail(
                'Code de réinitialisation de mot de passe',
                f'Votre code est : {code}',
                'ton-email@example.com',
                [email],
            )
            return redirect('verify_code')
    return render(request, 'auth/password_reset_request.html', context)


def verify_code(request):
    if request.method == "POST":
        form = VerificationCodeForm(request.POST)
        if form.is_valid():
            input_code = form.cleaned_data['code']
            saved_code = request.session.get('password_reset_code')
            code_time_str = request.session.get('password_reset_code_time')
            
            if not saved_code or not code_time_str:
                messages.error(request, "Aucun code trouvé. Veuillez recommencer.")
                return redirect('password_reset_request')
            
            code_time = timezone.datetime.fromisoformat(code_time_str)
           
            if timezone.is_naive(code_time):  # يعني إذا التاريخ بدون timezone
                 code_time = timezone.make_aware(code_time, timezone.get_current_timezone())
            now = timezone.now()
            diff_seconds = (now - code_time).total_seconds()
            
            if diff_seconds > 60:
                messages.error(request, "Le code de vérification a expiré. Veuillez demander un nouveau code.")
                return redirect('password_reset_request')
            
            if input_code == saved_code:
                # الكود صحيح، السماح بتغيير كلمة المرور
                request.session['password_reset_verified'] = True
                return redirect('set_new_password')
            else:
                messages.error(request, "Code invalide.")
    else:
        form = VerificationCodeForm()
    return render(request, 'auth/verify_code.html', {'form': form})

def set_new_password(request):
    if not request.session.get('password_reset_verified'):
        messages.error(request, "Vous devez vérifier votre code avant de changer le mot de passe.")
        return redirect('password_reset_request')

    if request.method == "POST":
        form = SetNewPasswordForm(request.POST)
        if form.is_valid():
            email = request.session.get('password_reset_email')
            if not email:
                messages.error(request, "Erreur interne, veuillez recommencer.")
                return redirect('password_reset_request')

            try:
                user = User.objects.get(email=email)
            except User.DoesNotExist:
                messages.error(request, "Utilisateur non trouvé.")
                return redirect('password_reset_request')

            user.set_password(form.cleaned_data['new_password'])
            user.save()

            # تنظيف الجلسة بعد نجاح العملية
            request.session.pop('password_reset_code', None)
            request.session.pop('password_reset_email', None)
            request.session.pop('password_reset_code_time', None)
            request.session.pop('password_reset_verified', None)

            messages.success(request, "Mot de passe changé avec succès.")
            return redirect('login')
    else:
        form = SetNewPasswordForm()
    
    return render(request, 'auth/set_new_password.html', {'form': form})

from django.http import HttpResponse
from django.template.loader import get_template
from xhtml2pdf import pisa

def pdf(request):
    equipments = Equipment.objects.all()
    template_path = 'pdf/equipment_pdf.html'
    context = {'equipments': equipments}
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="equipments.pdf"'
    template = get_template(template_path)
    html = template.render(context)

    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse('Erreur lors de la génération du PDF', status=500)
    return response
import openpyxl
from django.http import HttpResponse
from .models import Equipment  # تأكد من استيراد النموذج الصحيح

def export_excel(request):
    # جلب جميع المعدات (يمكن تعديل الاستعلام حسب حاجتك)
    equipments = Equipment.objects.all()

    # إنشاء ملف Excel جديد
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Equipments"

    # إضافة العناوين (header)
    ws.append(['Nom', 'Type', 'Numéro de Série', 'Date d\'achat'])

    # إضافة بيانات المعدات
    for eq in equipments:
        ws.append([eq.name, eq.type, eq.serial_number, eq.purchase_date.strftime("%d/%m/%Y")])

    # إعداد ملف الاستجابة للتحميل
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=equipment_list.xlsx'
    
    wb.save(response)
    return response
from django.contrib.auth import get_user_model
User = get_user_model()

import openpyxl
from django.http import HttpResponse
def export_xl(request):

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Utilisateurs"

    ws.append(['ID', 'Email', 'Nom', 'Prénom', 'Rôle', 'Actif'])

    users = User.objects.all()

    for user in users:
        ws.append([
            user.id,
            user.email,
            user.last_name,
            user.first_name,
            user.role if hasattr(user, 'role') else '',
            "Oui" if user.is_active else "Non",
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="utilisateurs.xlsx"'

    wb.save(response)
    return response
from django.http import HttpResponse
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from .models import CustomUser

def export_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="users.pdf"'

    pdf = canvas.Canvas(response, pagesize=letter)
    pdf.setFont("Helvetica", 12)  # Police standard

    users = CustomUser.objects.all()

    y = 750
    pdf.drawString(100, y, "Liste des utilisateurs")
    y -= 40

    for user in users:
        line = f"{user.id} - {user.email or ''} - {user.last_name or ''} - {user.first_name or ''} - {user.role or ''} - {'Oui' if user.is_active else 'Non'}"
        pdf.drawString(50, y, line)
        y -= 20
        if y < 50:
            pdf.showPage()
            pdf.setFont("Helvetica", 12)
            y = 750

    pdf.save()
    return response

from django.http import HttpResponse
from openpyxl import Workbook

from .models import CustomUser


def export_xl(request):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=utilisateurs.xlsx'

    wb = Workbook()
    ws = wb.active
    ws.title = "Utilisateurs"

    # En-têtes
    headers = ["ID", "Email", "Nom", "Prénom", "Rôle", "Actif", "Mot de passe"]
    ws.append(headers)

    users = CustomUser.objects.all()
    for user in users:
        role_display = {
            "admin": "Administrateur",
            "user": "Utilisateur"
        }.get(user.role, "Rôle inconnu")

        actif = "Oui" if user.is_active else "Non"

        row = [
            user.id,
            user.email,
            user.last_name,
            user.first_name,
            role_display,
            actif,
            user.plain_password if hasattr(user, 'plain_password') else "",
        ]
        ws.append(row)

    wb.save(response)
    return response
from django.shortcuts import get_object_or_404, redirect

def update_maintenance(request, pk):
    maintenance = get_object_or_404(Maintenance, pk=pk)
    if request.method == 'POST':
        maintenance.machine_id = request.POST.get('machine')
        maintenance.technicien_id = request.POST.get('technicien')
        maintenance.date = request.POST.get('date')
        maintenance.type = request.POST.get('type')
        maintenance.status = request.POST.get('status')
        maintenance.cost = request.POST.get('cost')
        maintenance.date_debut = request.POST.get('date_debut')
        maintenance.date_fin = request.POST.get('date_fin')
        maintenance.lieu_correction = request.POST.get('lieu_correction')
        maintenance.rapport = request.POST.get('rapport')
        if 'photo_piece' in request.FILES:
            maintenance.photo_piece = request.FILES['photo_piece']
        maintenance.save()
        return redirect('maintenance')  # Remplace par le nom de la vue principale
from django.http import HttpResponse
from django.template.loader import get_template
from xhtml2pdf import pisa

def export_maintenances_pdf(request):
    maintenances = Maintenance.objects.all()
    template = get_template('maintenance/pdf_template.html')
    html = template.render({'maintenances': maintenances})

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="maintenances.pdf"'

    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse('Une erreur est survenue lors de la génération du PDF', status=500)
    return response
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from .models import Maintenance

def export_maintenances_excel(request):
    # إنشاء مصنف وورقة عمل
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Maintenances"

    # عنوان الأعمدة
    headers = ["Machine", "Type", "État", "Cout", "Date de début", "Date de fin", "Technicien"]
    ws.append(headers)

    # استدعاء بيانات الصيانة
    maintenances = Maintenance.objects.all()

    # إضافة البيانات للورقة
    for m in maintenances:
        row = [
            m.machine.name if m.machine else '',
            m.type,
            m.status,
            f"{m.cost} DA" if m.cost is not None else '',
            m.date_debut.strftime('%d/%m/%Y') if m.date_debut else '',
            m.date_fin.strftime('%d/%m/%Y') if m.date_fin else '',
            m.technicien.email if m.technicien else '',  # تأكد أن technicien ليس None
        ]
        ws.append(row)

    # ضبط عرض الأعمدة تلقائيًا
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max_length + 2

    # إعداد استجابة HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=maintenances.xlsx'
    wb.save(response)

    return response
from rest_framework.generics import CreateAPIView

class PanneCreateAPIView(CreateAPIView):
    queryset = Panne.objects.all()
    serializer_class = PanneSerializer
    permission_classes = [IsAuthenticated]  # حسب الحاجة

    def perform_create(self, serializer):
        serializer.save(signalee_par=self.request.user)
from rest_framework.views import APIView 
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.db.models import Count, Avg, F, ExpressionWrapper, DurationField
from django.db.models.functions import TruncMonth
from django.utils.timezone import now
from datetime import timedelta, date
import datetime
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.db.models import Count, Avg, F, ExpressionWrapper, DurationField
from django.db.models.functions import TruncMonth
from django.utils.timezone import now
import datetime
from .models import Equipment, Panne, Maintenance

class DashboardDataAPIView(APIView):
    def get(self, request, format=None):
        try:
            start_date_str = request.GET.get('start_date')
            end_date_str = request.GET.get('end_date')

            start_date = None
            end_date = None

            if start_date_str:
                try:
                    start_date = datetime.date.fromisoformat(start_date_str)
                except ValueError:
                    return Response({'error': "start_date format invalide. Utilisez AAAA-MM-JJ."}, status=status.HTTP_400_BAD_REQUEST)
            if end_date_str:
                try:
                    end_date = datetime.date.fromisoformat(end_date_str)
                except ValueError:
                    return Response({'error': "end_date format invalide. Utilisez AAAA-MM-JJ."}, status=status.HTTP_400_BAD_REQUEST)

            equipment_qs = Equipment.objects.all()
            panne_qs = Panne.objects.all()
            maintenance_qs = Maintenance.objects.all()

            if start_date:
                panne_qs = panne_qs.filter(date_signalement__gte=start_date)
                maintenance_qs = maintenance_qs.filter(date_debut__gte=start_date)
            if end_date:
                panne_qs = panne_qs.filter(date_signalement__lte=end_date)
                maintenance_qs = maintenance_qs.filter(date_debut__lte=end_date)

            total_equipments = equipment_qs.count()
            broken_equipments = equipment_qs.filter(state='En Panne').count()

            if not start_date and not end_date:
                last_30_days = now().date() - datetime.timedelta(days=30)
                last_maintenance_count = Maintenance.objects.filter(date_debut__gte=last_30_days).count()
            else:
                last_maintenance_count = maintenance_qs.count()

            if panne_qs.exists():
                earliest_panne_date = panne_qs.earliest('date_signalement').date_signalement
                total_time = (now().date() - earliest_panne_date).days
            else:
                total_time = 0
            panne_count = panne_qs.count()
            mtbf = total_time / panne_count if panne_count else 0

            avg_duration = maintenance_qs.annotate(
                duration=ExpressionWrapper(
                    F('date_fin') - F('date_debut'),
                    output_field=DurationField()
                )
            ).aggregate(avg_hours=Avg('duration'))['avg_hours']
            avg_maintenance_hours = avg_duration.total_seconds() / 3600 if avg_duration else 0

            status_distribution = dict(equipment_qs.values_list('state').annotate(count=Count('id')))

            top_pannes = (panne_qs.values('equipment__name')
                          .annotate(count=Count('id'))
                          .order_by('-count')[:5])
            top_pannes_labels = [item['equipment__name'] for item in top_pannes]
            top_pannes_values = [item['count'] for item in top_pannes]

            panne_by_month = panne_qs.exclude(date_signalement__isnull=True).annotate(
                month=TruncMonth('date_signalement')
            ).values('month').annotate(count=Count('id')).order_by('month')

            maintenance_by_month = maintenance_qs.exclude(date_debut__isnull=True).annotate(
                month=TruncMonth('date_debut')
            ).values('month').annotate(count=Count('id')).order_by('month')

            monthly_chart = {
                'pannes': list(panne_by_month),
                'maintenances': list(maintenance_by_month),
            }

            timeline = []
            for panne in Panne.objects.select_related('equipment').order_by('-date_signalement')[:10]:
               maintenance = Maintenance.objects.filter(
               machine=panne.equipment,
               date_debut__gte=panne.date_signalement
            ).order_by('date_debut').first()

               timeline.append({
                  'equipment': panne.equipment.name,
                  'date_panne': panne.date_signalement.strftime('%d-%m-%Y'),
                  'date_maintenance': maintenance.date_fin.strftime('%d-%m-%Y') if maintenance and maintenance.date_fin else '—',
                  'duree_arret': f"{(maintenance.date_fin - panne.date_signalement).days} j"
                        if maintenance and maintenance.date_fin else '—',
                  'recommandation': "Vérifier les pièces critiques" if panne.niveau_urgence == 'élevé' else ''
                  })


            # Équipements nécessitant maintenance depuis > 60j
            equipements_needing_maintenance = []
            threshold_days = 60
            for eq in Equipment.objects.all():
                last_maintenance = Maintenance.objects.filter(machine=eq).order_by('-date_debut').first()
                if last_maintenance and last_maintenance.date_debut:
                    days_since = (now().date() - last_maintenance.date_debut).days
                    if days_since > threshold_days:
                        equipements_needing_maintenance.append({
                            'name': eq.name,
                            'days_since_last_maintenance': days_since
                        })

            # Équipements en maintenance depuis > 7j
            equipements_in_long_maintenance = []
            ongoing_maintenances = Maintenance.objects.filter(date_fin__isnull=True)
            for m in ongoing_maintenances:
                if m.date_debut:
                    days_in_maintenance = (now().date() - m.date_debut).days
                    if days_in_maintenance > 7:
                        equipements_in_long_maintenance.append({
                            'name': m.machine.name,
                            'days_in_maintenance': days_in_maintenance
                        })

            # Équipements sans maintenance depuis > 7j
            equipements_over_one_week_no_maintenance = []
            for eq in Equipment.objects.all():
                last_maintenance = Maintenance.objects.filter(machine=eq).order_by('-date_debut').first()
                if last_maintenance and last_maintenance.date_debut:
                    days_since = (now().date() - last_maintenance.date_debut).days
                    if days_since > 7:
                        equipements_over_one_week_no_maintenance.append({
                            'name': eq.name,
                            'days_since_last_maintenance': days_since
                        })

            return Response({
                'total_equipments': total_equipments,
                'broken_equipments': broken_equipments,
                'last_maintenance_count': last_maintenance_count,
                'mtbf_days': round(mtbf, 2),
                'avg_maintenance_hours': round(avg_maintenance_hours, 2),
                'status_distribution': status_distribution,
                'top_pannes': {
                    'labels': top_pannes_labels,
                    'data': top_pannes_values
                },
                'monthly_data': monthly_chart,
                'timeline': timeline,
                'recommandations': [
                    "Planifier une inspection générale",
                    "Analyser les pannes fréquentes",
                    "Remplacer les équipements trop anciens"
                ],
                'equipements_needing_maintenance': equipements_needing_maintenance,
                'equipements_in_long_maintenance': equipements_in_long_maintenance,
                'equipements_over_one_week_no_maintenance': equipements_over_one_week_no_maintenance,
            }, status=status.HTTP_200_OK)

        except Exception as e:
            print("❌ Erreur dans DashboardDataAPIView:", str(e))
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

from django.http import JsonResponse, HttpResponseNotAllowed
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status, viewsets
from .models import Notification, CustomUser, Group  # ajustez selon vos noms de modèles
from .serializers import NotificationSerializer, UserSerializer, GroupSerializer

# Afficher la page HTML principale
def notification_page(request):
    return render(request, 'notifications.html')

# API REST pour les notifications
@api_view(['GET', 'POST'])
def notification_list_create(request):
    if request.method == 'GET':
        notifications = Notification.objects.order_by('-created_at')
        serializer = NotificationSerializer(notifications, many=True)
        return Response({'results': serializer.data})

    elif request.method == 'POST':
        serializer = NotificationSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({'success': True, 'data': serializer.data}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

# Supprimer une notification
@api_view(['DELETE'])
def notification_delete(request, pk):
    notification = get_object_or_404(Notification, pk=pk)
    notification.delete()
    return JsonResponse({'success': True})

# Charger dynamiquement les utilisateurs
@api_view(['GET'])
def get_users(request):
    users = CustomUser.objects.all()
    serializer = UserSerializer(users, many=True)
    return Response(serializer.data)

# Charger dynamiquement les groupes
@api_view(['GET'])
def get_groups(request):
    groups = Group.objects.all()
    serializer = GroupSerializer(groups, many=True)
    return Response(serializer.data)
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
import json
from .models import Notification
from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group

User = get_user_model()

@csrf_exempt
def add_notification(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)

            message = data.get('message')
            type = data.get('type')
            recipient_type = data.get('recipient_type')
            user_email = data.get('user_email')
            group_id = data.get('group_id')

            if not message or not type or not recipient_type:
                return JsonResponse({'error': 'Champs requis manquants'}, status=400)

            notification = Notification(
                message=message,
                type=type,
            )

            # Affectation du destinataire
            if recipient_type == 'user' and user_email:
                try:
                    user = User.objects.get(email=user_email)
                    notification.recipient_user = user
                except User.DoesNotExist:
                    return JsonResponse({'error': "Utilisateur introuvable"}, status=400)
            elif recipient_type == 'group' and group_id:
                try:
                    group = Group.objects.get(id=group_id)
                    notification.recipient_group = group
                except Group.DoesNotExist:
                    return JsonResponse({'error': "Groupe introuvable"}, status=400)
            # Si 'all', on laisse les champs recipient_user et recipient_group vides (null)

            notification.save()
            return JsonResponse({'success': True, 'message': 'Notification ajoutée avec succès'})

        except json.JSONDecodeError:
            return JsonResponse({'error': 'Données JSON invalides'}, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    else:
        return JsonResponse({'error': 'Méthode non autorisée'}, status=405)
class UnreadNotificationCountAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        count = Notification.objects.filter(recipient_user=request.user, is_read=False).count()
        return Response({"count": count})

# views.py
from rest_framework.views import APIView
from rest_framework.response import Response
from .models import Notification
from .serializers import NotificationSerializer
from rest_framework.permissions import IsAuthenticated

class NotificationListAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        notifications = Notification.objects.filter(recipient_user=request.user).order_by('-created_at')
        serializer = NotificationSerializer(notifications, many=True)
        return Response(serializer.data)
from rest_framework.generics import CreateAPIView
from .models import Panne, Notification
from .serializers import PanneSerializer
from rest_framework.permissions import IsAuthenticated

class PanneCreateAPIView(CreateAPIView):
    queryset = Panne.objects.all()
    serializer_class = PanneSerializer
    permission_classes = [IsAuthenticated]

    def perform_create(self, serializer):
        panne = serializer.save(signalee_par=self.request.user)

        # Création automatique d'une notification liée à la panne
        Notification.objects.create(
            message=f"Nouvelle panne signalée sur l’équipement: {panne.equipment.name}",
            type="panne",  # ou un type personnalisé selon ton modèle
            recipient_user=None,  # Laisser vide si c'est une notification globale
        )
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from .models import Notification
from .serializers import NotificationSerializer

class NotificationListAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        notifications = Notification.objects.filter(
            recipient_user=request.user
        ).order_by('-created_at')
        serializer = NotificationSerializer(notifications, many=True)
        return Response(serializer.data)

from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status
from .models import Notification

@api_view(['POST'])
def mark_as_read(request):
    notif_id = request.data.get('id')
    try:
        notif = Notification.objects.get(id=notif_id)
        notif.is_read = True
        notif.save()
        return Response({"success": True})
    except Notification.DoesNotExist:
        return Response({"error": "Notification introuvable"}, status=status.HTTP_404_NOT_FOUND)
class MarkAllNotificationsAsReadView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        Notification.objects.filter(recipient_user=request.user, is_read=False).update(is_read=True)
        return Response({'message': 'Toutes les notifications ont été marquées comme lues.'})
# signals.py
from django.db.models.signals import post_save
from django.dispatch import receiver
from django.utils.timezone import now
from django.contrib.auth import get_user_model
from .models import Maintenance, Notification

User = get_user_model()

@receiver(post_save, sender=Maintenance)
def send_maintenance_notification(sender, instance, created, **kwargs):
    try:
        # فقط إذا status == terminee
        if instance.status == 'terminee':
            # تأكد أن الحالة تغيرت فعلاً
            if not created:
                old = Maintenance.objects.get(pk=instance.pk)
                if old.status == 'terminee':
                    return  # لا ترسل الإشعار مرتين

            # لكل مستخدم، أرسل إشعار لم يتم إرساله اليوم
            msg = f"L'équipement {instance.machine.name} a terminé sa maintenance et est retourné en fonctionnement."
            now_dt = now()
            for user in User.objects.all():
                deja = Notification.objects.filter(
                    recipient_user=user,
                    message=msg,
                    created_at__date=now_dt.date()
                ).exists()
                if not deja:
                    Notification.objects.create(
                        message=msg,
                        type='info',
                        recipient_user=user
                    )
    except Exception as e:
        print("Erreur dans le signal de maintenance:", e)
        
if not CustomUser.objects.filter(email='demo@example.com').exists():
    user = CustomUser.objects.create_user(
        email='demo@example.com',
        password='demo1234',
        first_name='User',
        last_name='Demo',
        role='admin'  # أو 'technician' أو حسب الحقول المطلوبة في نموذجك
    )
    user.is_staff = True
    user.is_superuser = True
    user.save()
    print("✅ Demo user created successfully!")
